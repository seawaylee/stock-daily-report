from datetime import datetime

def save_to_excel(df, filename="results/fish_basin_report.xlsx"):
    """
    Save the dataframe to an Excel file with conditional formatting.
    Red for Bullish/Positive, Green for Bearish/Negative.
    """
    if df.empty: return
    
    # Ensure directory exists
    import os
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    
    # 1. Create a Styler
    # Helper to color text
    def color_status(val):
        color = 'red' if val == 'YES' else 'green'
        return f'color: {color}'
        
    def color_pct(val):
        # val is string like "+1.23%" or "-0.5%"
        try:
            v = float(val.strip('%'))
            color = 'red' if v > 0 else 'green'
            return f'color: {color}'
        except:
            return ''

    # Apply styles
    # Note: subset needs columns to exist
    try:
        # 动态检测可用的百分比列
        pct_columns = [c for c in ['涨幅%', '黄线偏离率', '白线偏离率', '偏离率', '区间涨幅%'] if c in df.columns]
        styler = df.style.map(color_status, subset=['状态'])\
                        .map(color_pct, subset=pct_columns)

        # 高亮条件: 金叉=1天、死叉=1天、或状态转换日期是今天
        def highlight_important_row(row):
            today_str = datetime.now().strftime("%y.%m.%d")
            should_highlight = False
            
            # 检查金叉天数=1
            if '金叉天数' in row and row['金叉天数'] == 1:
                should_highlight = True
            # 检查死叉天数=1
            if '死叉天数' in row and row['死叉天数'] == 1:
                should_highlight = True
            # 检查状态变量时间是否是今天
            if '状态变量时间' in row:
                status_time = str(row['状态变量时间']).strip()
                if status_time == today_str:
                    should_highlight = True
            
            if should_highlight:
                return ['background-color: #FFFFCC'] * len(row)  # Light Yellow
            return [''] * len(row)

        styler = styler.apply(highlight_important_row, axis=1)

                         
        # 2. Save
        styler.to_excel(filename, index=False, engine='openpyxl')
        print(f"Excel report saved to: {filename}")
        
        # 3. Post-process for column widths - 优化中文字符宽度
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        ws = wb.active
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    # 中文字符按2个字符宽度计算
                    length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                    if length > max_length:
                        max_length = length
                except:
                    pass
            # 确保最小宽度为8，加上padding
            adjusted_width = max(max_length + 3, 8)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        wb.save(filename)
        
    except Exception as e:
        print(f"Failed to save Excel: {e}")


def merge_excel_sheets(index_path, sector_path, output_path):
    """
    合并指数和题材的Excel到一个sheet中，中间用分隔行隔开
    """
    import os
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    
    try:
        # 读取两个Excel
        if not os.path.exists(index_path) or not os.path.exists(sector_path):
            print(f"源文件不存在: {index_path} 或 {sector_path}")
            return False
            
        df_index = pd.read_excel(index_path)
        df_sector = pd.read_excel(sector_path)
        
        # 确保输出目录存在
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # 创建分隔行 DataFrame
        separator_data = {col: [''] for col in df_index.columns}
        separator_data[df_index.columns[0]] = ['═══════════ 题材趋势 ═══════════']
        df_separator = pd.DataFrame(separator_data)
        
        # 创建指数标题行
        index_title_data = {col: [''] for col in df_index.columns}
        index_title_data[df_index.columns[0]] = ['═══════════ 指数趋势 ═══════════']
        df_index_title = pd.DataFrame(index_title_data)
        
        # 合并: 指数标题 + 指数数据 + 空行 + 题材标题 + 题材数据
        # 确保题材列与指数列对齐
        for col in df_index.columns:
            if col not in df_sector.columns:
                df_sector[col] = ''
        df_sector = df_sector[df_index.columns]
        
        df_combined = pd.concat([
            df_index_title,
            df_index,
            df_separator,
            df_sector
        ], ignore_index=True)
        
        # 重新应用样式
        def color_status(val):
            if not isinstance(val, str): return ''
            if val == 'YES': return 'color: red'
            if val == 'NO': return 'color: green'
            return ''
            
        def color_pct(val):
            try:
                if not isinstance(val, str): return ''
                v = float(val.strip('%'))
                return 'color: red' if v > 0 else 'color: green'
            except: return ''

        pct_columns = [c for c in ['涨幅%', '黄线偏离率', '白线偏离率', '偏离率', '区间涨幅%'] if c in df_combined.columns]
        
        styler = df_combined.style.map(color_status, subset=['状态'])\
                                .map(color_pct, subset=pct_columns)
        
        # 高亮逻辑
        def highlight_important_row(row):
            # 跳过分隔行
            if '═══════════' in str(row.iloc[0]):
                return [''] * len(row)
                
            today_str = datetime.now().strftime("%y.%m.%d")
            should_highlight = False
            
            try:
                # 检查金叉天数=1
                if '金叉天数' in row and str(row['金叉天数']) == '1':
                    should_highlight = True
                # 检查死叉天数=1
                if '死叉天数' in row and str(row['死叉天数']) == '1':
                    should_highlight = True
                # 检查状态变量时间是否是今天
                if '状态变量时间' in row:
                    status_time = str(row['状态变量时间']).strip()
                    if status_time == today_str:
                        should_highlight = True
            except: pass
            
            if should_highlight:
                return ['background-color: #FFFFCC'] * len(row)
            return [''] * len(row)

        styler = styler.apply(highlight_important_row, axis=1)

        # 保存带样式的Excel
        styler.to_excel(output_path, index=False, engine='openpyxl')
        
        # 二次处理：调整列宽和分隔行样式
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 调整列宽
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_name = str(column[0].value) if column[0].value else ""
            column_cells = [cell for cell in column]
            
            for cell in column_cells:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    # 跳过分隔行的大标题计算，避免把它算进第一列的宽度
                    if '═══════════' in cell_value:
                        continue
                        
                    length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                    if length > max_length:
                        max_length = length
                except:
                    pass
            
            # 基础宽度
            adjusted_width = max(max_length + 2, 8)
            
            # 特殊列处理
            if column_name == '代码':
                adjusted_width = min(adjusted_width, 12) # 代码列不应该超过12
            elif column_name in ['状态', '涨幅%', '排名变化', '金叉天数', '死叉天数']:
                adjusted_width = min(adjusted_width, 10) # 这些列紧凑一点
                
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
        
        # 高亮分隔行并居中
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        bold_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center', vertical='center')
        
        for row in ws.iter_rows():
            cell_value = str(row[0].value) if row[0].value else ""
            if '═══════════' in cell_value:
                # 找到该行最后一个有数据的列
                max_col = ws.max_column
                # 合并单元格 (A列到最后一列)
                start_cell = row[0]
                end_cell = row[max_col-1]
                ws.merge_cells(start_row=start_cell.row, start_column=start_cell.column, 
                               end_row=end_cell.row, end_column=end_cell.column)
                
                # 应用样式到合并后的单元格
                start_cell.fill = yellow_fill
                start_cell.font = bold_font
                start_cell.alignment = center_align
                
                # 同时也给合并区域内的其他单元格上色，不然会有白边
                for cell in row:
                    cell.fill = yellow_fill
            else:
                # 普通数据行，统一居中对齐
                for cell in row:
                   cell.alignment = center_align
        
        wb.save(output_path)
        print(f"✅ 合并Excel已保存: {output_path}")
        return True
        
    except Exception as e:
        print(f"合并Excel失败: {e}")
        import traceback
        traceback.print_exc()
        return False

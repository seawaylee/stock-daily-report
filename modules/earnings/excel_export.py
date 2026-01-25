
"""
Excel Export for Earnings Analysis (Using openpyxl)
"""
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

def export_earnings_excel(df, output_path):
    """
    Export analyzed earnings data to Excel with formatting using openpyxl.
    """
    try:
        # Select and Rename Columns
        output_cols = [
            '股票代码', '股票简称', '实际披露时间', 
            '预告类型', '预测指标', '业绩变动幅度', '预测数值', 
            'sentiment', 'analysis_summary', '业绩变动原因'
        ]
        
        # Ensure sentiment/summary columns exist
        if 'sentiment' not in df.columns:
            df['sentiment'] = '中性'
        if 'analysis_summary' not in df.columns:
            df['analysis_summary'] = ''
            
        # Filter existing columns
        valid_cols = [c for c in output_cols if c in df.columns]
        final_df = df[valid_cols].copy()
        
        # Rename for display
        column_map = {
            '股票代码': '代码',
            '股票简称': '名称',
            '实际披露时间': '日期',
            'sentiment': '评估',
            'analysis_summary': '核心观点'
        }
        final_df = final_df.rename(columns=column_map)
        
        # Use openpyxl engine
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            sheet_name = '业绩披露分析'
            final_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Access workbook/worksheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Styles
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header_font = Font(bold=True)
            
            # Good (Red text, light red bg)
            good_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            good_font = Font(color="9C0006")
            
            # Bad (Green text, light green bg) 
            bad_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            bad_font = Font(color="006100")
            
            # Neutral (Yellow)
            neutral_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            neutral_font = Font(color="9C6500")

            # 1. Format Headers
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                
            # 2. Adjust Column Widths
            for i, column_cells in enumerate(worksheet.columns, 1):
                length = max(len(str(cell.value) or "") for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(i)].width = min(length + 5, 50)
                
            # 3. Conditional Formatting on '评估' column
            # Find column index for '评估'
            eval_col_idx = None
            for idx, col_name in enumerate(final_df.columns):
                if col_name == '评估':
                    eval_col_idx = idx + 1 # 1-based
                    break
            
            if eval_col_idx:
                col_letter = get_column_letter(eval_col_idx)
                cell_range = f"{col_letter}2:{col_letter}{len(final_df)+1}"
                
                # Good
                dxf_good = DifferentialStyle(font=good_font, fill=good_fill)
                rule_good = Rule(type="containsText", operator="containsText", text="利好", dxf=dxf_good)
                rule_good.formula = [f'NOT(ISERROR(SEARCH("利好",{col_letter}2)))']
                worksheet.conditional_formatting.add(cell_range, rule_good)
                
                # Bad
                dxf_bad = DifferentialStyle(font=bad_font, fill=bad_fill)
                rule_bad = Rule(type="containsText", operator="containsText", text="利空", dxf=dxf_bad)
                rule_bad.formula = [f'NOT(ISERROR(SEARCH("利空",{col_letter}2)))']
                worksheet.conditional_formatting.add(cell_range, rule_bad)
                
                # Neutral
                dxf_neu = DifferentialStyle(font=neutral_font, fill=neutral_fill)
                rule_neu = Rule(type="containsText", operator="containsText", text="中性", dxf=dxf_neu)
                rule_neu.formula = [f'NOT(ISERROR(SEARCH("中性",{col_letter}2)))']
                worksheet.conditional_formatting.add(cell_range, rule_neu)

        return True
    except Exception as e:
        print(f"Excel Export Failed: {e}")
        import traceback
        traceback.print_exc()
        return False
